import os
import sys
import logging

import time
import base64
import requests
import io
import json
import re
import unicodedata
import psycopg2
import psycopg2.extras
from datetime import datetime
from dateutil import parser

from flask import render_template, request, jsonify, redirect, url_for, flash, send_file
from werkzeug.utils import secure_filename

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Side, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference

from app.utils.helpers import get_departments, get_zone_data, allowed_file
from blacklist.blacklist_tracker import BlacklistTracker
from models.db import get_transaksi_filtered, get_grafik_data

# ─── Logging Setup ─────────────────────────────────────────────────────────────
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        logging.FileHandler("app.log"),
        logging.StreamHandler(sys.stdout)
    ]
)

logger = logging.getLogger(__name__)

def get_zona_from_device(device_name, hijau_key, merah_key):
    hijau_devices = [d.strip().lower() for d in os.getenv(hijau_key, "").split(",")]
    merah_devices = [d.strip().lower() for d in os.getenv(merah_key, "").split(",")]

    name = (device_name or "").lower()
    hijau = name if any(h in name for h in hijau_devices) else ""
    merah = name if any(m in name for m in merah_devices) else ""
    return hijau, merah

def apply_excel_header(ws, tahun: int):
    for col in "ABCDEFGHIJKLMN":
        ws.column_dimensions[col].auto_size = True
    ws.column_dimensions['M'].width = 18
    ws.column_dimensions['N'].width = 35

    ws.merge_cells('A1:N5')
    ws.merge_cells('A6:B7')
    ws.merge_cells('A8:B8')
    ws.merge_cells('A9:B9')
    ws.merge_cells('C6:D7')
    ws.merge_cells('C8:D8')
    ws.merge_cells('C9:D9')
    ws.merge_cells('E6:L9')
    ws.merge_cells('M6:M6')
    ws.merge_cells('M7:M7')
    ws.merge_cells('M8:M8')
    ws.merge_cells('M9:M9')
    ws.merge_cells('N6:N6')
    ws.merge_cells('N7:N7')
    ws.merge_cells('N8:N8')
    ws.merge_cells('N9:N9')
    ws.merge_cells('A10:A11')
    ws.merge_cells('B10:B11')
    ws.merge_cells('C10:C11')
    ws.merge_cells('D10:D11')
    ws.merge_cells('E10:E11')
    ws.merge_cells('F10:F11')
    ws.merge_cells('G10:G11')
    ws.merge_cells('H10:I10')
    ws.merge_cells('J10:J11')
    ws.merge_cells('K10:L10')
    ws.merge_cells('M10:M11')
    ws.merge_cells('N10:N11')

    ws["A6"] = "HARI"
    ws["A8"] = "TANGGAL"
    ws["A9"] = "JAM"
    ws["E6"] = os.getenv("EXCEL_TITLE", "MONITORING MASUK KELUAR ORANG PT. PLN INDONESIA POWER UBP SURALAYA")
    ws["M6"] = "NOMOR DOKUMEN"
    ws["N6"] = os.getenv("NOMOR_DOKUMEN", "PB.13.7.20.7.7FRM.18.SLA")
    ws["M7"] = "TANGGAL"
    ws["N7"] = f"15 MEI {tahun}"
    ws["M8"] = "REVISI"
    ws["N8"] = "00"
    ws["M9"] = "HALAMAN"

    ws["A10"] = "NO"
    ws["B10"] = "NAMA PERUSAHAAN"
    ws["C10"] = "NAMA PEGAWAI"
    ws["D10"] = "NIP"
    ws["E10"] = "JABATAN"
    ws["F10"] = "JENIS PEKERJAAN"
    ws["G10"] = "FIRST IN TIME"
    ws["H10"] = "ZONA MASUK"
    ws["H11"] = "AREA TERBATAS"
    ws["I11"] = "AREA TERLARANG"
    ws["J10"] = "LAST OUT TIME"
    ws["K10"] = "ZONA KELUAR"
    ws["K11"] = "AREA TERBATAS"
    ws["L11"] = "AREA TERLARANG"
    ws["M10"] = "NO. SPK/LOI/MEMO"
    ws["N10"] = "PARAF / NAMA ANGGOTA SATPAM"

    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    for row in ws.iter_rows(min_row=6, max_row=11, min_col=1, max_col=14):
        for cell in row:
            cell.alignment = align_center
            cell.border = border

def normalize(text: str) -> str:
    if not text:
        return ""
    # Ubah ke huruf besar, hapus karakter non-alfanumerik, ubah spasi ke _
    text = unicodedata.normalize("NFKD", text).encode("ASCII", "ignore").decode("utf-8")
    text = text.upper()
    text = re.sub(r"[^\w\s]", "", text)
    text = re.sub(r"\s+", "_", text)
    return text.strip()

def match_device_name(device_name: str, device_pos_list: list[str]) -> str:
    if not device_name:
        return ""

    norm_device_name = normalize(device_name)
    normalized_pos_map = {normalize(pos): pos for pos in device_pos_list if pos}

    for norm_pos, original_pos in normalized_pos_map.items():
        if norm_pos in norm_device_name:
            # logging.info(f"[MATCH] Device '{device_name}' cocok dengan ENV '{original_pos}'")
            return original_pos

    # Coba cari pola "POS" + angka
    match_pos_number = re.search(r'pos[\s_\-]*([0-9]+)', device_name, re.IGNORECASE)
    if match_pos_number:
        number = match_pos_number.group(1)
        label = f"POS {int(number)}"
        # logging.info(f"[MATCH] Device '{device_name}' dikenali sebagai '{label}' dari angka")
        return label

    # Cari pola "POS" + label
    match_pos_label = re.search(r'pos[\s_\-]+([a-zA-Z]+)', device_name, re.IGNORECASE)
    if match_pos_label:
        label = match_pos_label.group(1).upper()
        result = f"POS {label}"
        # logging.info(f"[MATCH] Device '{device_name}' dikenali sebagai '{result}' dari label")
        return result

    # Gagal cocokkan
    # logging.warning(f"[NO MATCH] Device '{device_name}' tidak cocok dengan ENV atau pola")
    return device_name

def get_attribute_values(conn, person_id: str, attr_names: list[str]) -> dict:
    """
    Mengembalikan dictionary: {attr_name: value}, berdasarkan filed_index dari pers_attribute.
    Ambil data dari pers_attribute_ext berdasarkan kolom bernama attr_value{index}.
    """
    attr_names = [attr.strip().upper() for attr in attr_names if attr.strip()]
    if not attr_names:
        return {}

    filed_indexes = {}
    with conn.cursor() as cursor:
        # Step 1: Ambil filed_index untuk setiap attr_name dari pers_attribute
        for attr in attr_names:
            cursor.execute("""
                SELECT filed_index
                FROM pers_attribute
                WHERE UPPER(attr_name) = %s
                LIMIT 1
            """, (attr,))
            res = cursor.fetchone()
            if res and res[0] is not None:
                filed_indexes[attr] = res[0]

        if not filed_indexes:
            return {}

        # Step 2: Ambil kolom attr_valueN sesuai filed_index
        column_names = [f"attr_value{index}" for index in filed_indexes.values()]
        sql_columns = ", ".join(column_names)

        cursor.execute(f"""
            SELECT {sql_columns}
            FROM pers_attribute_ext
            WHERE person_id = %s
            LIMIT 1
        """, (person_id,))
        row = cursor.fetchone()

    if not row:
        return {}

    attr_values = {}
    for i, (attr, index) in enumerate(filed_indexes.items()):
        col_name = f"attr_value{index}"
        value = row[i] if i < len(row) else None
        attr_values[attr] = value if value is not None else ""

    return attr_values

def write_excel_data(ws, records, conn):
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=False)
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # Ambil atribut dari ENV dan pastikan huruf besar
    attr_names = [a.strip().upper() for a in os.getenv("ATTRIBUT_TRANSAKSI", "").split(",") if a.strip()]
    device_pos_list = [x.strip() for x in os.getenv("DEVICE_POS", "").split(",") if x.strip()]

    row_num = 12
    no = 1

    for record in records:
        dept_name = record.get("dept_name", "")
        name = record.get("name", "")
        pin = record.get("pin", "")
        person_id = record.get("id", "")

        first_in = record["first_in_time"].strftime("%Y-%m-%d %H:%M:%S") if record.get("first_in_time") else ""
        last_out = record["last_out_time"].strftime("%Y-%m-%d %H:%M:%S") if record.get("last_out_time") else ""

        # Nama device dari DB
        device_in_raw = record.get("reader_name_in", "")
        device_out_raw = record.get("reader_name_out", "")

        # Zona hijau/merah
        hijauin, merahin = get_zona_from_device(device_in_raw, "IN_DEVICES_HIJAU", "IN_DEVICES_MERAH")
        hijauout, merahout = get_zona_from_device(device_out_raw, "OUT_DEVICES_HIJAU", "OUT_DEVICES_MERAH")

        # Pemetaan nama device
        device_in_hijau = match_device_name(hijauin, device_pos_list)
        device_in_merah = match_device_name(merahin, device_pos_list)
        device_out_hijau = match_device_name(hijauout, device_pos_list)
        device_out_merah = match_device_name(merahout, device_pos_list)

        # Ambil data tambahan dari pers_attribute_ext
        attr_values = get_attribute_values(conn, person_id, attr_names)

        # Ambil nilai-nilai spesifik dari attr_values sesuai urutan
        nip = attr_values.get(attr_names[0], "") if len(attr_names) > 0 else ""
        jenis_pekerjaan = attr_values.get(attr_names[1], "") if len(attr_names) > 1 else ""
        jabatan = attr_values.get(attr_names[2], "") if len(attr_names) > 2 else ""
        no_po = attr_values.get(attr_names[3], "") if len(attr_names) > 3 else ""

        # Tulis ke worksheet
        ws.append([
            no,
            dept_name, name, nip,
            jabatan, jenis_pekerjaan, first_in,
            device_in_hijau, device_in_merah,
            last_out,
            device_out_hijau, device_out_merah,
            no_po, ""
        ])

        # Format cell
        for cell in next(ws.iter_rows(min_row=row_num, max_row=row_num, min_col=1, max_col=14)):
            cell.alignment = align_center
            cell.border = border

        row_num += 1
        no += 1

    return row_num

def auto_adjust_column_width(ws, start_row=12, min_width=20):
    column_widths = {}

    for row in ws.iter_rows(min_row=start_row, values_only=True):
        for i, cell_value in enumerate(row, 1):
            if cell_value is None:
                continue
            length = len(str(cell_value))
            if i not in column_widths:
                column_widths[i] = length
            else:
                if length > column_widths[i]:
                    column_widths[i] = length

    max_col = ws.max_column
    for col_idx in range(1, max_col + 1):
        width = column_widths.get(col_idx, 0)
        final_width = max(width + 2, min_width)
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = final_width

def register_routes(app):
    token = os.getenv("ACCESS_TOKEN")
    url_add = os.getenv("URL_ADD_PERSON")
    title_hijau = os.getenv("TITLE_HIJAU", "MONITORING ZONA HIJAU")
    title_merah = os.getenv("TITLE_MERAH", "MONITORING ZONA MERAH")
    title_all = os.getenv("TITLE_ALL", "MONITORING SEMUA ZONA")
    transaksi_title = os.getenv("TRANSAKSI_TITLE", "Riwayat Transaksi PLN Indonesia Power")

    def get_conn():
        return psycopg2.connect(dsn=os.getenv("DATABASE_URL"))

    ZONA_HIJAU = [z.strip().lower() for z in os.getenv("ZONA_HIJAU", "").split(",")]
    ZONA_MERAH = [z.strip().lower() for z in os.getenv("ZONA_MERAH", "").split(",")]

    def is_zona(name, zone_list):
        name = (name or "").lower()
        return any(z in name for z in zone_list)

    # == Tambahkan route ini di bawah semua route lain ==
    @app.route('/export')
    def export():
        conn = get_conn()
        cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

        from_date = request.args.get("from")
        to_date = request.args.get("to")
        nama = request.args.get("nama", "")
        dept = request.args.get("dept", "")
        pin = request.args.get("id", "")

        if not from_date or not to_date:
            return {"error": "Parameter 'from' dan 'to' harus diisi."}, 400

        query = "SELECT * FROM acc_firstin_lastout WHERE update_time BETWEEN %s AND %s"
        params = [from_date, to_date]

        if pin:
            query += " AND pin ILIKE %s"
            params.append(f"%{pin}%")
        if nama:
            query += " AND name ILIKE %s"
            params.append(f"%{nama}%")
        if dept:
            query += " AND dept_name ILIKE %s"
            params.append(f"%{dept}%")

        query += " ORDER BY first_in_time NULLS LAST"
        cur.execute(query, tuple(params))
        records = cur.fetchall()

        wb = Workbook()
        ws = wb.active
        tahun = datetime.now().year

        if not records:
            logger.info(f"Export kosong dari {request.remote_addr}: {from_date} - {to_date}, filter={nama or pin or dept}")
            return {"error": "Data tidak ditemukan dalam rentang waktu tersebut."}, 404

        logger.info(f"Export berhasil oleh {request.remote_addr}: {len(records)} data dari {from_date} ke {to_date}, filter={nama or pin or dept}")

        apply_excel_header(ws, tahun)
        write_excel_data(ws, records, conn)
        auto_adjust_column_width(ws)

        ip_logo_path = os.getenv("EXCEL_LOGO_KIRI")
        ipp_logo_path = os.getenv("EXCEL_LOGO_KANAN")

        if ip_logo_path and os.path.exists(ip_logo_path):
            img = XLImage(ip_logo_path)
            img.height = 50
            img.anchor = 'A2'
            ws.add_image(img)

        if ipp_logo_path and os.path.exists(ipp_logo_path):
            img2 = XLImage(ipp_logo_path)
            img2.height = 40
            img2.anchor = 'H2'
            ws.add_image(img2)

        virtual_file = io.BytesIO()

        try:
            wb.save(virtual_file)
        except Exception as e:
            logger.error(f"Gagal menyimpan file Excel: {e}")
            return {"error": "Gagal menyimpan file."}, 500

        virtual_file.seek(0)
        file_name = f"transaction_plnn_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return send_file(
            virtual_file,
            as_attachment=True,
            download_name=file_name,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )


    @app.context_processor
    def inject_now():
        return {'current_year': datetime.now().year}

    # ─── Error Handler ─────────────────
    @app.errorhandler(404)
    def page_not_found(e):
        return render_template("errors/404.html"), 404

    @app.errorhandler(500)
    def internal_server_error(e):
        return render_template("errors/500.html"), 500

    # ─── Zona View ─────────────────────
    @app.route("/")
    def zona_hijau():
        return render_template("index.html", title=title_hijau, zone="hijau")

    @app.route("/merah")
    def zona_merah():
        return render_template("index.html", title=title_merah, zone="merah")

    @app.route("/all")
    def zona_all():
        return render_template("all/index.html", title=title_all, zone="all")

    @app.route("/transaksi")
    def transaksi():
        return render_template("transaksi/index.html", title=transaksi_title)

    @app.route("/grafik")
    def grafik():
        grafik_title = os.getenv("GRAFIK_TITLE", "Grafik Keluar Masuk Per Zona")
        return render_template("grafik/index.html", title=grafik_title)

    # ─── API Zone Data ─────────────────
    @app.route("/api/data")
    def api_data():
        return jsonify(get_zone_data("hijau"))

    @app.route("/api/merah")
    def api_merah():
        return jsonify(get_zone_data("merah"))

    @app.route("/api/blacklist")
    def api_blacklist():
        return jsonify(BlacklistTracker().run())

    @app.route("/api/all")
    def api_all():
        data_hijau = get_zone_data("hijau")
        data_merah = get_zone_data("merah")

        return jsonify({
            "hijau": data_hijau,
            "merah": data_merah
        })

    @app.route("/api/transaksi")
    def api_transaksi():
        id_ = request.args.get("id", "")
        nama = request.args.get("nama", "")
        dept = request.args.get("dept", "")
        dari = request.args.get("dari", "")
        ke = request.args.get("ke", "")
        page = int(request.args.get("page", 1))
        per_page = 50

        if not dari or not ke:
            now = datetime.now()
            dari = now.strftime("%Y-%m-%dT00:00:00")
            ke = now.strftime("%Y-%m-%dT23:59:59")

        result, total = get_transaksi_filtered(id_, nama, dept, dari, ke, page, per_page)

        return jsonify({
            "total": total,
            "per_page": per_page,
            "page": page,
            "rows": result
        })

    def _process_zone_events(events, in_devices, out_devices):
        """
        Synchronous replica of EventProcessor state-machine logic for one zone.
        Returns a list of logical in/out events after per-person deduplication.
        """
        in_devs = {d.strip().upper() for d in in_devices}
        out_devs = {d.strip().upper() for d in out_devices}
        reader_in = {d for d in in_devs if "-READER" in d}
        reader_out = {d for d in out_devs if "-READER" in d}

        def get_type(dev_name):
            dev_upper = dev_name.strip().upper()
            if dev_upper in {d.replace("-READER", "") for d in in_devs}:
                return "in"
            if dev_upper in {d.replace("-READER", "") for d in out_devs}:
                return "out"
            return None

        def ts_from_val(val):
            try:
                if isinstance(val, datetime):
                    return int(val.timestamp())
                return int(datetime.strptime(str(val).strip(), "%Y-%m-%d %H:%M:%S").timestamp())
            except (ValueError, TypeError, OSError):
                return None

        # Step 1: collect events per person (same as EventProcessor)
        per_person = {}
        for e in events:
            pin = (e.get("pin") or "").strip()
            name = (e.get("name") or "").strip()
            dept = (e.get("dept_name") or "").strip() or "UNKNOWN"

            dev_alias = str(e.get("dev_alias") or "").strip().upper()
            event_point_name = str(e.get("event_point_name") or "").strip().upper()

            dev = dev_alias
            for reader_dev in (reader_in | reader_out):
                base_name = reader_dev.replace("-READER", "").strip().upper()
                if event_point_name == base_name:
                    dev = event_point_name
                    break

            time_raw = e.get("event_time")
            ts = ts_from_val(time_raw)
            if not pin or not dev or ts is None:
                continue

            ev_type = get_type(dev)
            if not ev_type:
                continue

            person = per_person.setdefault(pin, {"dept": dept, "name": name, "events": []})
            person["events"].append({"type": ev_type, "ts": ts, "time": time_raw})

        # Step 2: state-machine per person (same logic as EventProcessor)
        logical_events = []
        for pin, person in per_person.items():
            events_sorted = sorted(person["events"], key=lambda x: x["ts"])
            status = "outside"

            for ev in events_sorted:
                if ev["type"] == "in" and status == "outside":
                    status = "inside"
                    logical_events.append({
                        "dept": person["dept"],
                        "action": "in",
                        "time": ev["time"],
                    })
                elif ev["type"] == "out" and status == "inside":
                    status = "outside"
                    logical_events.append({
                        "dept": person["dept"],
                        "action": "out",
                        "time": ev["time"],
                    })

        return logical_events

    def _aggregate_grafik(dari, ke, mode):
        """
        Aggregate entry/exit/current data per period, department, and zone
        using the same EventProcessor state-machine logic as zone pages.
        """
        empty = {
            "labels": [],
            "pos1_in": [], "pos1_out": [], "pos1_cur": [],
            "pos2_in": [], "pos2_out": [], "pos2_cur": [],
            "departments": []
        }
        try:
            events = get_grafik_data(dari, ke)
        except Exception as e:
            logger.error(f"Gagal mengambil data grafik: {e}")
            return empty

        if not events:
            return empty

        # Device configs (same env vars as tracker_worker)
        in_hijau = [d.strip() for d in os.getenv("IN_DEVICES_HIJAU", "").split(",") if d.strip()]
        out_hijau = [d.strip() for d in os.getenv("OUT_DEVICES_HIJAU", "").split(",") if d.strip()]
        in_merah = [d.strip() for d in os.getenv("IN_DEVICES_MERAH", "").split(",") if d.strip()]
        out_merah = [d.strip() for d in os.getenv("OUT_DEVICES_MERAH", "").split(",") if d.strip()]

        # Process each zone through the state machine
        pos1_events = _process_zone_events(events, in_hijau, out_hijau)
        pos2_events = _process_zone_events(events, in_merah, out_merah)

        def get_period_label(time_val):
            try:
                if isinstance(time_val, datetime):
                    dt = time_val
                else:
                    dt = datetime.strptime(str(time_val).strip(), "%Y-%m-%d %H:%M:%S")
            except (ValueError, TypeError):
                return "UNKNOWN"
            if mode == "week":
                iso = dt.isocalendar()
                return f"{iso[0]}-W{iso[1]:02d}"
            elif mode == "month":
                return dt.strftime("%Y-%m")
            else:
                return dt.strftime("%Y-%m-%d")

        period_agg = {}
        dept_agg = {}

        def add_event(zone_prefix, ev):
            label = get_period_label(ev["time"])
            dept = ev["dept"]
            action = ev["action"]
            key = f"{zone_prefix}_{action}"

            if label not in period_agg:
                period_agg[label] = {
                    "pos1_in": 0, "pos1_out": 0,
                    "pos2_in": 0, "pos2_out": 0,
                }
            period_agg[label][key] += 1

            if dept not in dept_agg:
                dept_agg[dept] = {
                    "pos1_in": 0, "pos1_out": 0,
                    "pos2_in": 0, "pos2_out": 0,
                }
            dept_agg[dept][key] += 1

        for ev in pos1_events:
            add_event("pos1", ev)
        for ev in pos2_events:
            add_event("pos2", ev)

        # current (di dalam) = in - out, same as SummaryBuilder
        for p in period_agg.values():
            p["pos1_cur"] = max(p["pos1_in"] - p["pos1_out"], 0)
            p["pos2_cur"] = max(p["pos2_in"] - p["pos2_out"], 0)

        sorted_labels = sorted(period_agg.keys())

        departments = []
        for dept in sorted(dept_agg.keys()):
            d = dept_agg[dept]
            departments.append({
                "dept": dept,
                "pos1_in": d["pos1_in"], "pos1_out": d["pos1_out"],
                "pos1_cur": max(d["pos1_in"] - d["pos1_out"], 0),
                "pos2_in": d["pos2_in"], "pos2_out": d["pos2_out"],
                "pos2_cur": max(d["pos2_in"] - d["pos2_out"], 0),
            })

        return {
            "labels": sorted_labels,
            "pos1_in": [period_agg[l]["pos1_in"] for l in sorted_labels],
            "pos1_out": [period_agg[l]["pos1_out"] for l in sorted_labels],
            "pos1_cur": [period_agg[l]["pos1_cur"] for l in sorted_labels],
            "pos2_in": [period_agg[l]["pos2_in"] for l in sorted_labels],
            "pos2_out": [period_agg[l]["pos2_out"] for l in sorted_labels],
            "pos2_cur": [period_agg[l]["pos2_cur"] for l in sorted_labels],
            "departments": departments,
        }

    @app.route("/api/grafik")
    def api_grafik():
        dari = request.args.get("dari", "")
        ke = request.args.get("ke", "")
        mode = request.args.get("mode", "day")

        if not dari or not ke:
            now = datetime.now()
            dari = now.strftime("%Y-%m-%dT00:00:00")
            ke = now.strftime("%Y-%m-%dT23:59:59")

        return jsonify(_aggregate_grafik(dari, ke, mode))

    @app.route("/export_grafik")
    def export_grafik():
        dari = request.args.get("dari", "")
        ke = request.args.get("ke", "")
        mode = request.args.get("mode", "day")

        if not dari or not ke:
            return {"error": "Parameter 'dari' dan 'ke' harus diisi."}, 400

        data = _aggregate_grafik(dari, ke, mode)

        if not data["labels"]:
            return {"error": "Data tidak ditemukan dalam rentang waktu tersebut."}, 404

        mode_labels = {"day": "Per Hari", "week": "Per Minggu", "month": "Per Bulan"}

        wb = Workbook()

        # ── Sheet 1: Per-department summary ──
        ws_dept = wb.active
        ws_dept.title = "Data Per Departemen"

        ws_dept.merge_cells("A1:H1")
        ws_dept["A1"] = f"Grafik Keluar Masuk Per Zona — {mode_labels.get(mode, mode)}"
        ws_dept["A1"].font = Font(bold=True, size=14)
        ws_dept["A1"].alignment = Alignment(horizontal="center", vertical="center")

        ws_dept.merge_cells("A2:H2")
        ws_dept["A2"] = f"Periode: {dari} s/d {ke}"
        ws_dept["A2"].alignment = Alignment(horizontal="center")

        header_row = 4
        dept_headers = ["DEPARTEMEN", "POS 1 MASUK", "POS 1 KELUAR", "POS 1 DI DALAM",
                        "POS 2 MASUK", "POS 2 KELUAR", "POS 2 DI DALAM"]
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin")
        )

        for col_idx, h in enumerate(dept_headers, 1):
            cell = ws_dept.cell(row=header_row, column=col_idx, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = align_center
            cell.border = border

        departments = data.get("departments", [])
        tot_p1i = tot_p1o = tot_p1c = tot_p2i = tot_p2o = tot_p2c = 0
        for i, dept in enumerate(departments):
            row_num = header_row + 1 + i
            values = [
                dept["dept"],
                dept["pos1_in"], dept["pos1_out"], dept["pos1_cur"],
                dept["pos2_in"], dept["pos2_out"], dept["pos2_cur"],
            ]
            tot_p1i += dept["pos1_in"]
            tot_p1o += dept["pos1_out"]
            tot_p1c += dept["pos1_cur"]
            tot_p2i += dept["pos2_in"]
            tot_p2o += dept["pos2_out"]
            tot_p2c += dept["pos2_cur"]
            for col_idx, val in enumerate(values, 1):
                cell = ws_dept.cell(row=row_num, column=col_idx, value=val)
                cell.alignment = align_center
                cell.border = border

        # Total row
        total_row = header_row + 1 + len(departments)
        total_vals = ["TOTAL", tot_p1i, tot_p1o, tot_p1c, tot_p2i, tot_p2o, tot_p2c]
        for col_idx, val in enumerate(total_vals, 1):
            cell = ws_dept.cell(row=total_row, column=col_idx, value=val)
            cell.font = Font(bold=True)
            cell.alignment = align_center
            cell.border = border

        for col_idx in range(1, 8):
            ws_dept.column_dimensions[get_column_letter(col_idx)].width = 20

        # Department chart
        if departments:
            chart_dept = BarChart()
            chart_dept.type = "col"
            chart_dept.grouping = "clustered"
            chart_dept.title = "Data Per Departemen"
            chart_dept.y_axis.title = "Jumlah Orang"
            chart_dept.x_axis.title = "Departemen"
            chart_dept.width = 30
            chart_dept.height = 14

            chart_data_ref = Reference(ws_dept, min_col=2, min_row=header_row, max_col=7, max_row=total_row - 1)
            cats_ref = Reference(ws_dept, min_col=1, min_row=header_row + 1, max_row=total_row - 1)
            chart_dept.add_data(chart_data_ref, titles_from_data=True)
            chart_dept.set_categories(cats_ref)

            dept_colors = ["28A745", "90EE90", "198754", "DC3545", "FF6B6B", "B02A37"]
            for idx, color in enumerate(dept_colors):
                if idx < len(chart_dept.series):
                    chart_dept.series[idx].graphicalProperties.solidFill = color

            ws_dept.add_chart(chart_dept, f"A{total_row + 2}")

        # ── Sheet 2: Per-period time series ──
        ws_ts = wb.create_sheet("Data Per Periode")

        ws_ts.merge_cells("A1:H1")
        ws_ts["A1"] = f"Grafik Keluar Masuk Per Zona — {mode_labels.get(mode, mode)}"
        ws_ts["A1"].font = Font(bold=True, size=14)
        ws_ts["A1"].alignment = Alignment(horizontal="center", vertical="center")

        ws_ts.merge_cells("A2:H2")
        ws_ts["A2"] = f"Periode: {dari} s/d {ke}"
        ws_ts["A2"].alignment = Alignment(horizontal="center")

        ts_header_row = 4
        ts_headers = ["Periode", "POS 1 Masuk", "POS 1 Keluar", "POS 1 Di Dalam",
                       "POS 2 Masuk", "POS 2 Keluar", "POS 2 Di Dalam"]
        for col_idx, h in enumerate(ts_headers, 1):
            cell = ws_ts.cell(row=ts_header_row, column=col_idx, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = align_center
            cell.border = border

        for i, label in enumerate(data["labels"]):
            row_num = ts_header_row + 1 + i
            values = [
                label,
                data["pos1_in"][i], data["pos1_out"][i], data["pos1_cur"][i],
                data["pos2_in"][i], data["pos2_out"][i], data["pos2_cur"][i],
            ]
            for col_idx, val in enumerate(values, 1):
                cell = ws_ts.cell(row=row_num, column=col_idx, value=val)
                cell.alignment = align_center
                cell.border = border

        last_ts_row = ts_header_row + len(data["labels"])

        for col_idx in range(1, 8):
            ws_ts.column_dimensions[get_column_letter(col_idx)].width = 18

        # Time-series chart
        chart = BarChart()
        chart.type = "col"
        chart.grouping = "clustered"
        chart.title = f"Grafik Keluar Masuk — {mode_labels.get(mode, mode)}"
        chart.y_axis.title = "Jumlah Orang"
        chart.x_axis.title = "Periode"
        chart.width = 30
        chart.height = 14

        chart_data = Reference(ws_ts, min_col=2, min_row=ts_header_row, max_col=7, max_row=last_ts_row)
        cats = Reference(ws_ts, min_col=1, min_row=ts_header_row + 1, max_row=last_ts_row)
        chart.add_data(chart_data, titles_from_data=True)
        chart.set_categories(cats)

        ts_colors = ["28A745", "90EE90", "198754", "DC3545", "FF6B6B", "B02A37"]
        for idx, color in enumerate(ts_colors):
            if idx < len(chart.series):
                chart.series[idx].graphicalProperties.solidFill = color

        ws_ts.add_chart(chart, f"A{last_ts_row + 2}")

        logger.info(f"Export grafik oleh {request.remote_addr}: {dari} - {ke}, mode={mode}, {len(data['labels'])} periode, {len(departments)} dept")

        virtual_file = io.BytesIO()
        try:
            wb.save(virtual_file)
        except Exception as e:
            logger.error(f"Gagal menyimpan file Excel grafik: {e}")
            return {"error": "Gagal menyimpan file."}, 500

        virtual_file.seek(0)
        file_name = f"grafik_{mode}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return send_file(
            virtual_file,
            as_attachment=True,
            download_name=file_name,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    @app.route("/search_person")
    def search_person():
        term = request.args.get("q", "").strip()
        if not term:
            return jsonify([])

        conn = get_conn()
        cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)

        cur.execute("""
            SELECT p.name AS person_name, p.pin, d.name AS dept_name
            FROM pers_person p
            JOIN auth_department d ON p.auth_dept_id = d.id
            WHERE p.name ILIKE %s OR CAST(p.pin AS TEXT) ILIKE %s
            ORDER BY p.name
            LIMIT 10
        """, (f"%{term}%", f"%{term}%"))

        rows = cur.fetchall()
        cur.close()
        conn.close()

        results = [
            {
                "name": row["person_name"],
                "pin": row["pin"],
                "dept_name": row["dept_name"]
            }
            for row in rows
        ]
        return jsonify(results)


    @app.route("/register_visitor", methods=["GET", "POST"])
    def register_visitor():
        if request.method == "POST":
            try:
                certNum = request.form.get("certNum", "").strip()
                company = request.form.get("company", "").strip()
                startTime = request.form.get("startTime", "").strip()
                endTime = request.form.get("endTime", "").strip()
                persPersonPin = request.form.get("persPersonPin", "").strip()  # hasil dari select2
                visEmpName = request.form.get("visEmpName", "").strip()
                # visitEmpPhone = request.form.get("visitEmpPhone", "").strip()
                visitReason = request.form.get("visitReason", "").strip()

                file = request.files.get("facePhoto")
                if not file or not allowed_file(file.filename):
                    flash("File foto tidak valid (hanya .jpg/.png)", "danger")
                    return redirect(url_for("register_visitor"))

                filename = secure_filename(file.filename)
                filepath = os.path.join(app.config["UPLOAD_FOLDER"], filename)
                file.save(filepath)

                with open(filepath, "rb") as f:
                    encoded_photo = base64.b64encode(f.read()).decode()

                payload = {
                    "cardNo": "",
                    "certNum": certNum,
                    "certType": 1,
                    "company": company,
                    "startTime": startTime,
                    "endTime": endTime,
                    "persPersonPin": persPersonPin,
                    "visEmpName": visEmpName,
                    "visLevels": "1",
                    "visitEmpPhone": "",
                    "visitReason": visitReason,
                    "visitorCount": 1,
                    "facePhoto": encoded_photo
                }

                headers = {
                    "Content-Type": "application/json",
                    "Accept": "application/json"
                }

                url_visitor = f"https://localhost:8098/api/visRegistration/add?access_token={token}"

                response = requests.post(
                    url_visitor,
                    json=payload,
                    headers=headers,
                    verify=False,
                    timeout=10
                )

                data = response.json()
                msg = data.get("message", "") or data.get("status", "")

                if msg.lower() == "success":
                    flash("Registrasi visitor berhasil", "success")
                else:
                    flash(f"Registrasi visitor gagal: {msg}", "danger")

            except Exception:
                flash("Terjadi kesalahan saat memproses data visitor", "danger")
                app.logger.exception("Register visitor error:")

            return redirect(url_for("register_visitor"))

        return render_template("register_visitor.html")

    # ─── Register Form ─────────────────
    @app.route("/register", methods=["GET", "POST"])
    def register():
        departments = get_departments()
        if not departments:
            return render_template("register.html", offline=True)

        if request.method == "POST":
            try:
                name = request.form.get("name", "").strip().upper()
                dept = request.form.get("dept", "").strip().upper()
                plat = request.form.get("plat", "").strip().upper()
                gender = request.form.get("gender", "M")

                file = request.files.get("filename")
                if not file or not allowed_file(file.filename):
                    flash("File tidak valid (hanya .jpg/.png)", "danger")
                    return redirect(url_for("register"))

                filename = secure_filename(file.filename)
                filepath = os.path.join(app.config["UPLOAD_FOLDER"], filename)
                file.save(filepath)

                with open(filepath, "rb") as f:
                    encoded = base64.b64encode(f.read()).decode()

                pin = str(int(time.time() * 1000))[-8:]

                payload = {
                    "name": name,
                    "pin": pin,
                    "deptCode": dept,
                    "gender": gender,
                    "carPlate": plat,
                    "personPhoto": encoded,
                    "accLevelIds": "1",
                    "certType": 2,
                    "ssn": "111111",
                    "isDisabled": False,
                    "isSendMail": False
                }

                headers = {
                    "Content-Type": "application/json",
                    "Accept": "application/json"
                }

                response = requests.post(
                    f"{url_add}?access_token={token}",
                    json=payload,
                    headers=headers,
                    verify=False,
                    timeout=10
                )

                data = response.json()
                msg = data.get("message", "Gagal")

                if msg == "success":
                    flash("Registrasi berhasil", "success")

                    attr_names = os.getenv("ATTRIBUT_REGISTER", "").split(",")
                    attr_names = [a.strip() for a in attr_names if a.strip()]

                    if attr_names:
                        conn = get_conn()
                        cur = conn.cursor()

                        cur.execute("SELECT id FROM pers_person WHERE pin = %s LIMIT 1", (pin,))
                        person = cur.fetchone()
                        if not person:
                            app.logger.warning(f"Person ID dengan PIN {pin} tidak ditemukan.")
                            cur.close()
                            conn.close()
                            return redirect(url_for("register"))

                        person_id = person[0]

                        # Ambil attr_name dan filed_index dari tabel pers_attribute
                        cur.execute("SELECT attr_name, filed_index FROM pers_attribute")
                        attr_map = {row[0].strip().upper(): row[1] for row in cur.fetchall() if row[0]}

                        for attr_name in attr_names:
                            form_key = attr_name.lower().replace(" ", "_")
                            value = request.form.get(form_key, "").strip()
                            if not value:
                                continue

                            filed_index = attr_map.get(attr_name.upper())
                            if filed_index is None:
                                app.logger.warning(f"Attribute '{attr_name}' tidak ditemukan di pers_attribute.")
                                continue

                            column_name = f"attr_value{filed_index}"

                            update_query = f"""
                                UPDATE pers_attribute_ext
                                SET "{column_name}" = %s
                                WHERE person_id = %s
                            """
                            cur.execute(update_query, (value, person_id))

                        conn.commit()
                        cur.close()
                        conn.close()

                else:
                    flash(f"Registrasi gagal: {msg}", "danger")

            except Exception:
                flash("Terjadi kesalahan saat memproses data", "danger")
                app.logger.exception("Register error:")

            return redirect(url_for("register"))

        # GET: render form + extra fields
        extra_fields = [f.strip().upper() for f in os.getenv("ATTRIBUT_REGISTER", "").split(",") if f.strip()]
        return render_template("register.html", departments=departments, extra_fields=extra_fields)